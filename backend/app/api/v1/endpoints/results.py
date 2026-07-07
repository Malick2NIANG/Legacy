"""
Endpoints de recuperation et d export des resultats d experiences.
Fournit les metriques, la matrice de confusion et les exports CSV/JSON.
"""
import csv
import io
import json
from datetime import timezone
from zoneinfo import ZoneInfo
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
from app.api.deps import get_db, get_current_user
from app.models.result import Result
from app.models.experiment import Experiment
from app.models.dataset import Dataset
from app.models.model import Model
from app.services.storage_service import StorageService

DAKAR = ZoneInfo("Africa/Dakar")


def _fmt_dakar(dt) -> str:
    """Formate un datetime en heure de Dakar (UTC+0, Africa/Dakar)."""
    if dt is None:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(DAKAR).strftime("%Y-%m-%d %H:%M:%S")

router = APIRouter()


class ResultRead(BaseModel):
    id: int
    experiment_id: int
    accuracy: Optional[float]
    precision: Optional[float]
    recall: Optional[float]
    f1_score: Optional[float]
    confusion_matrix: Optional[list]
    training_history: Optional[dict]
    model_key: Optional[str] = None

    class Config:
        from_attributes = True


def _get_result_or_404(experiment_id: int, owner_id: int, db: Session) -> Result:
    exp = db.query(Experiment).filter(Experiment.id == experiment_id).first()
    if not exp or exp.owner_id != owner_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Experience introuvable")
    result = db.query(Result).filter(Result.experiment_id == experiment_id).first()
    if not result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Resultats non disponibles - experience pas encore terminee",
        )
    return result


@router.get("/{experiment_id}/download-model")
def download_model(
    experiment_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Stream le fichier modele (.pkl/.h5/.pt) depuis MinIO — pas de presigned URL."""
    from fastapi.responses import StreamingResponse
    result = _get_result_or_404(experiment_id, current_user.id, db)
    if not result.model_key:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Modele non disponible — relancez l'experience pour generer le fichier.",
        )
    storage = StorageService()
    try:
        stream   = storage.get_model_stream(result.model_key)
        filename = result.model_key.split("/")[-1]
        return StreamingResponse(
            stream,
            media_type="application/octet-stream",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur de téléchargement : {e}")


@router.get("/{experiment_id}", response_model=ResultRead)
def get_results(
    experiment_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Retourne les metriques completes d une experience terminee."""
    return _get_result_or_404(experiment_id, current_user.id, db)


@router.get("/{experiment_id}/export")
def export_results(
    experiment_id: int,
    format: str = Query("json", regex="^(json|csv|csv_history)$"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Exporte les resultats.
    - format=csv          : resume complet (1 ligne) avec metadonnees + metriques
    - format=csv_history  : historique d entrainement epoch par epoch
    - format=json         : tout en JSON
    """
    result = _get_result_or_404(experiment_id, current_user.id, db)

    exp = db.query(Experiment).filter(Experiment.id == experiment_id).first()
    dataset = db.query(Dataset).filter(Dataset.id == exp.dataset_id).first() if exp else None
    model = db.query(Model).filter(Model.id == exp.model_id).first() if exp else None

    duration_s = None
    if exp and exp.created_at and exp.finished_at:
        duration_s = int((exp.finished_at - exp.created_at).total_seconds())

    cm = result.confusion_matrix or []
    cm_flat = {}
    if cm:
        for i, row in enumerate(cm):
            for j, val in enumerate(row):
                cm_flat[f"cm_{i}_{j}"] = val

    history = result.training_history or {}

    if format == "csv":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter  # noqa: F401

        # ── Couleurs Legacy ─────────────────────────────────────
        GREEN      = "00853F"
        DARK       = "0D2818"
        GOLD       = "F5A227"
        LIGHT_GREEN= "E6F4ED"
        LIGHT_GOLD = "FEF3E2"
        WHITE      = "FFFFFF"
        GRAY       = "F4F7F5"

        def _fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def _border():
            s = Side(style="thin", color="D6E8DC")
            return Border(left=s, right=s, top=s, bottom=s)

        wb = Workbook()
        ws = wb.active
        ws.title = "Résultats Legacy"

        # ── Titre principal ─────────────────────────────────────
        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        title_cell.value = "LEGACY — Rapport de Résultats"
        title_cell.font      = Font(bold=True, size=14, color=WHITE)
        title_cell.fill      = _fill(DARK)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.append([])  # ligne vide

        # ── Statut ──────────────────────────────────────────────
        status_str = ""
        if exp and exp.status:
            s = str(exp.status)
            status_str = s.split(".")[-1] if "." in s else s

        # ── Section helper ──────────────────────────────────────
        def _section(label):
            ws.append([label])
            row = ws.max_row
            cell = ws.cell(row=row, column=1)
            cell.font      = Font(bold=True, size=11, color=WHITE)
            cell.fill      = _fill(GREEN)
            cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[row].height = 20
            ws.merge_cells(f"A{row}:C{row}")

        def _row(key, value, fill_color=None):
            ws.append([key, value])
            row = ws.max_row
            for col in [1, 2]:
                c = ws.cell(row=row, column=col)
                c.border = _border()
                if col == 1:
                    c.font = Font(bold=True, size=10, color=DARK)
                    c.fill = _fill(GRAY)
                else:
                    c.font = Font(size=10, color=DARK)
                    c.fill = _fill(fill_color or WHITE)

        # ── Informations expérience ─────────────────────────────
        _section("INFORMATIONS EXPÉRIENCE")
        _row("ID",          experiment_id)
        _row("Nom",         exp.name if exp else "")
        _row("Dataset",     dataset.name if dataset else "")
        _row("Modèle",      model.name if model else "")
        _row("Type",        model.model_type if model else "")
        _row("Statut",      status_str)
        _row("Créé le",     _fmt_dakar(exp.created_at) if exp else "")
        _row("Terminé le",  _fmt_dakar(exp.finished_at) if exp else "")
        if duration_s is not None:
            mins, secs = divmod(duration_s, 60)
            _row("Durée",   f"{mins}m {secs}s")
        else:
            _row("Durée",   "")
        ws.append([])

        # ── Métriques ───────────────────────────────────────────
        _section("MÉTRIQUES")
        metrics = [
            ("Accuracy",  result.accuracy),
            ("Precision", result.precision),
            ("Recall",    result.recall),
            ("F1 Score",  result.f1_score),
        ]
        for name, val in metrics:
            pct = round(val * 100, 2) if val is not None else None
            fill = LIGHT_GREEN if pct and pct >= 80 else (LIGHT_GOLD if pct and pct >= 60 else WHITE)
            _row(name, f"{pct} %" if pct is not None else "", fill_color=fill)
        ws.append([])

        # ── Matrice de confusion ────────────────────────────────
        if cm:
            _section("MATRICE DE CONFUSION")
            n = len(cm)
            header_row = [""] + [f"Prédit {i}" for i in range(n)]
            ws.append(header_row)
            hrow = ws.max_row
            for col_idx in range(1, n + 2):
                c = ws.cell(row=hrow, column=col_idx)
                c.font = Font(bold=True, color=WHITE, size=10)
                c.fill = _fill(GOLD)
                c.alignment = Alignment(horizontal="center")
                c.border = _border()

            for i, row_data in enumerate(cm):
                ws.append([f"Réel {i}"] + list(row_data))
                r = ws.max_row
                ws.cell(row=r, column=1).font = Font(bold=True, color=DARK, size=10)
                ws.cell(row=r, column=1).fill = _fill(GRAY)
                ws.cell(row=r, column=1).border = _border()
                for j, val in enumerate(row_data):
                    c = ws.cell(row=r, column=j + 2)
                    # Diagonale = bon résultat → vert
                    c.fill = _fill(LIGHT_GREEN if i == j else WHITE)
                    c.font = Font(bold=(i == j), color=DARK, size=10)
                    c.alignment = Alignment(horizontal="center")
                    c.border = _border()

        # ── Largeurs colonnes ────────────────────────────────────
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 16

        # ── Export ──────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"legacy_resultats_{experiment_id}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if format == "csv_history":
        if not history:
            raise HTTPException(status_code=404, detail="Historique d entrainement non disponible")
        keys = list(history.keys())
        n_epochs = len(history[keys[0]]) if keys else 0
        output = io.StringIO()
        fieldnames = ["epoch"] + keys
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        for epoch in range(n_epochs):
            row = {"epoch": epoch + 1}
            for k in keys:
                v = history[k][epoch]
                row[k] = round(v, 6) if isinstance(v, float) else v
            writer.writerow(row)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=history_{experiment_id}.csv"},
        )

    data = {
        "experiment": {
            "id":          experiment_id,
            "name":        exp.name if exp else None,
            "dataset":     dataset.name if dataset else None,
            "model":       model.name if model else None,
            "model_type":  model.model_type if model else None,
            "status":      str(exp.status) if exp else None,
            "created_at":  _fmt_dakar(exp.created_at) if exp and exp.created_at else None,
            "finished_at": _fmt_dakar(exp.finished_at) if exp and exp.finished_at else None,
            "duration_s":  duration_s,
        },
        "metrics": {
            "accuracy":  result.accuracy,
            "precision": result.precision,
            "recall":    result.recall,
            "f1_score":  result.f1_score,
        },
        "confusion_matrix": cm,
        "training_history": history,
    }
    return StreamingResponse(
        io.BytesIO(json.dumps(data, indent=2).encode()),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=results_{experiment_id}.json"},
    )
